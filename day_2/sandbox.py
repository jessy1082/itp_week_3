#import os
import openpyxl
from openpyxl import workbook
from openpyxl.workbook.workbook import Workbook

wb = openpyxl.load_workbook('C:/Users/jessy/Desktop/Vetstech_Python/itp_week_3/day_2/lecture.xlsx')
#workbook = openpyxl.load_workbook('day_2/lecture.xlsx')
#print(str(wb.sheetnames))

#data = workbook.sheet.cell()
my_sheet1 = wb['Population by Census Tract']
my_column = my_sheet1

for i in range(1, my_sheet1.max_row +1):
   # print(i)
    if i == None:
       break
    i = str(i)
    print(i+ "Im a string")