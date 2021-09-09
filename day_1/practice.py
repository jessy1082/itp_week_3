import openpyxl

wb = openpyxl.load_workbook("C:/Users/jessy/Desktop/Vetstech_Python/itp_week_3/day_1/lecture.xlsx")
print(type(wb))

sheet= wb['Sheet1']
print(sheet.max_row +1) # R: 7

# for i in range(1, sheet.max_row +1):
for i in range(1,8):
    # on the date of A, C amount of Bwere sold
    date= "A" + str(i)
    date_cell = sheet[date]

    amount = 'C' + str(i)
    amount_cell = sheet[amount]

    fruit = "B" + str(i)
    fruit_cell = sheet[fruit]

    print ("On the date of " + str(date_cell.value) + "," + 
    str(amount_cell.value) + " amount of " + fruit_cell.value + "!")