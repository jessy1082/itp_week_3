# ITP Week 3 Day 1 Lecture

#------------PIP-----------
# C - create
# R - read
# U - update
# D - delete

#------------EXCEL SPREADSHEET-----------

# -Workbook
#     -can contain multiple sheets (worksheets)
# -xlsx file extension
# -Sheets  / Worksheets
# -Columns (letters) & Rows (numbers)
# -Cell
#     -intersection of row & Column

# READ
# - SHEET, CELLS, COLUMNS, ROWS

#------------OPENPYXL-----------
# The OpenPyXL third-party module handles Excel spreadsheets (.xlsx files).

# -openpyxl is a downloaded module
# -Download & install openpyxl using pip
#   pip3 install openpyxl (on gitbash)

import openpyxl

#Assign your .xlsx file to a variable
filename = "C:/Users/jessy/Desktop/Vetstech_Python/itp_week_3/day_1/lecture.xlsx"
wb = openpyxl.load_workbook(filename) #returns a Workbook object.
print(type(wb)) #<class 'openpyxl.workbook.workbook.Workbook'>  
# #my_workbook = openpyxl.load_workbook('example.xlsx')
# #use the 'type' method on the new variable to verify what kind of data type you are working with
# print(type(wb)) #<class 'openpyxl.workbook.workbook.Workbook'>
#type(my_workboo)


sheet_name = wb.sheetnames #property of specific objest
print(sheet_name)


# sheet_name = wb.sheetnames
# print(sheet_name) #['Sheet1', 'Population by Census Tract']

# #Each Excel file can have multiple sheets, which each contain many rows, columns and cells.  You must specify which sheet you would like to manipulate.
# #get_sheet_names() #and
# #get_sheet_by_name() #help get Worksheet objects.

# #my_workbook.get_sheet_names()  # Result-->  ['Sheet1', 'Sheet2', 'Sheet3']

my_sheet = wb['Sheet1'] #key is the sheetnames itself

#my_sheet = wb.get_sheet_by_name('Sheet1')
#wb.get_sheet_by_name(sheet_name[0])  to get sheet name

# my_sheet = wb['Sheet1']
# print(my_sheet) # R: <Worksheet "Sheet1">


# # my_sheet = my_workbook.get_sheet_by_name('Sheet 1')
# # verify
# # type(my_sheet)  # Result--> R: <class 'openpyxl.worksheet.worksheet.Worksheet'>

apple_cell = my_sheet['B1'] #object
print(apple_cell.value) # R: Apples

another_cell = my_sheet['B3'] 
print(another_cell.value) # R: Pears

print('Row %s, column %s, is %s' %(another_cell.row, another_cell.column, another_cell.value))

print('COORDINATE %s is %s' %(another_cell.coordinate, another_cell.value))

# # Cell objects have a "value" member variable with the content of that cell.
# # The square brackets in sheet['A1'] get Cell objects.
# # The cell() method also returns a Cell object from a sheet.
# #my_sheet['A1'] #will return the CELL OBJECT (NOT the value you see in the cell) created by the intersection of the first column and the first row.  The reult of sheet['A1'] will actually be:
#     # Result -->  <Cell Sheet1.A1>
# #so we must ASSIGN IT TO A VARIABLE before we can work with it.
# #my_cell = sheet['A1']
# #my_cell.value() # Result --> the value of the A1 cell

# #Don't forget to convert the values to a string if you are working with strings
# #str(my_cell.value)
# #str(my_sheet['A1'].value)

# #You can also access cell values using the .cell() method and passing arguments for the column and row:
# #my_sheet.cell(row=1, column=2) # Result --> <Cell Sheet1.B1>




# #Python can be used to loop through the cells on your spreadsheet and manipulate data automatically.
# #for i in range(1, 8):
#  #   print(i, my_sheet.cell(row = i, column =2).value)
# #The code may be difficult to understand at first, but notice that in 'row = i', the 'i' will increase by 1 until it reaches 8, thus printing out each value in column 2.  The first 'i', after the 'print(' is to show the correlation between 'i' and row being printed.


# # READ DATA

# # WRITE TO A DATA

# # UPDATE A SPREADSHEET

# # SET UP A DATA STRUCTURE WITH THE UPDATED INFORMATION

# # FORMULAS

# # OTHER FUNCTIONALITIES

# # SUMMARY


